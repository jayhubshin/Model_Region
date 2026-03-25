import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter
import io
import time
import re
from datetime import datetime, date
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import pytz
import folium
from streamlit_folium import st_folium
from folium.plugins import MarkerCluster
from urllib.parse import quote

# 페이지 설정
st.set_page_config(
    page_title="충전기 모델분류 자동화",
    page_icon="?",
    layout="wide"
)

def get_korea_time():
    """한국 시간(KST) 반환"""
    korea_tz = pytz.timezone('Asia/Seoul')
    return datetime.now(korea_tz)

def get_safe_value(row_data, col_letter):
    """엑셀 셀 값을 안전하게 문자열로 가져오는 헬퍼 함수"""
    val = row_data.get(col_letter)
    if val is None:
        return ""
    return str(val).strip()

def format_time(seconds):
    """초를 읽기 쉬운 형식으로 변환"""
    if seconds < 60:
        return f"{seconds:.1f}초"
    elif seconds < 3600:
        minutes = int(seconds // 60)
        secs = int(seconds % 60)
        return f"{minutes}분 {secs}초"
    else:
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        return f"{hours}시간 {minutes}분"

def clean_and_parse_date(date_value):
    """날짜 데이터 정리 및 변환"""
    if date_value is None:
        return None
    
    if isinstance(date_value, datetime):
        return date_value.date()
    if isinstance(date_value, date):
        return date_value
    
    if isinstance(date_value, str):
        date_str = date_value.strip()
        
        # 시간 부분 제거
        date_str = re.sub(r'\s+\d{2}:\d{2}:\d{2}$', '', date_str)
        date_str = re.sub(r'\s+00:00:00$', '', date_str)
        date_str = re.sub(r'\s+0:00:00$', '', date_str)
        
        if not date_str:
            return None
        
        date_formats = [
            '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d',
            '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S',
            '%d-%m-%Y', '%d/%m/%d',
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
    
    try:
        if isinstance(date_value, (int, float)):
            from datetime import timedelta
            excel_epoch = datetime(1899, 12, 30)
            return (excel_epoch + timedelta(days=float(date_value))).date()
    except:
        pass
    
    return None

def format_date_for_excel(date_obj):
    """date 객체를 엑셀용 YYYY-MM-DD 문자열로 변환"""
    if date_obj is None:
        return None
    if isinstance(date_obj, date):
        return date_obj.strftime('%Y-%m-%d')
    return None

def classify_region(address):
    """H열의 주소 데이터를 기반으로 권역을 분류하는 함수"""
    if not address:
        return "기타"
    
    address_clean = str(address).strip()
    
    # 인천을 먼저 검사
    if re.search(r"인천", address_clean):
        if re.search(r"계양|남동|동구|미추홀|부평|연수|서구|중구|강화", address_clean):
            return "수도권남서"
        else:
            return "인천??"
    
    # 서울/경기 검사
    elif re.search(r"서울|경기", address_clean):
        if re.search(r"고양|부천|김포|파주|은평구|마포구|서대문구|양천구|강서구|용산구|중구|종로구", address_clean):
            return "수도권북서"
        elif re.search(r"도봉구|노원구|중랑구|강북구|성북구|동대문구|성동구|광진구|의정부|남양주|구리|양주|포천|동두천|가평|연천", address_clean):
            return "수도권북동"
        elif re.search(r"강남구|서초구|송파구|강동구|성남|용인|하남|광주|안성|수원|평택|오산|이천|여주|양평", address_clean):
            return "수도권남동"
        elif re.search(r"구로구|금천구|영등포구|동작구|관악구|의왕|광명|군포|과천|시흥|안산|안양|화성", address_clean):
            return "수도권남서"
        else:
            return "수도권기타"
    
    # 광역권 분류
    elif re.search(r"강원", address_clean):
        return "강원권"
    elif re.search(r"충청|충남|충북|세종|대전", address_clean):
        return "충청권"
    elif re.search(r"경상|경남|경북|부산|대구|울산", address_clean):
        return "경상권"
    elif re.search(r"전라|전남|전북|광주", address_clean):
        return "전라권"
    else:
        return "기타"

def classify_model(row_data, row_num):
    """엑셀 IFS 수식을 파이썬 로직으로 변환한 모델 분류 함수"""
    AD = get_safe_value(row_data, 'AD')
    AG = get_safe_value(row_data, 'AG')
    AH = get_safe_value(row_data, 'AH')
    AJ = get_safe_value(row_data, 'AJ')
    
    if AH == "급속":
        ag_left4 = AG[:4] if len(AG) >= 4 else AG
        
        if ag_left4 == "S0F1":
            return "급속스필_100"
        elif ag_left4 == "S0F5":
            return "급속스필_50"
        elif ag_left4 == "EVQ-" and AJ == "100":
            return "급속PNE_100"
        elif (ag_left4 == "EVQ-" or ag_left4 == "EV1-") and AJ == "50":
            return "급속PNE_50"
        elif ag_left4 == "MAXE":
            return "급속PNE_200"
        elif ag_left4 == "DP15":
            return "급속PNE_150"
        elif ag_left4 in ["A01-", "AD1-"]:
            return "급속애플망고_200"
        elif ag_left4 in ["Q081", "Q101", "Q010"]:
            return "급속SK_100"
        elif ag_left4 in ["Q071", "Q102"]:
            return "급속SK_200"
        elif ag_left4 in ["1Y25", "1Y24"]:
            return "급속코스텔_50"
        elif ag_left4 == "1911":
            return "급속중앙제어_50"
        elif ag_left4 == "1900":
            return "급속그린파워_100"
        elif ag_left4 == "19C0":
            return "급속그린파워_50"
        elif ag_left4 == "QC50":
            return "급속알박_50"
        else:
            return "급속"
    
    ag_left3 = AG[:3] if len(AG) >= 3 else AG
    ag_left4 = AG[:4] if len(AG) >= 4 else AG
    ag_left6 = AG[:6] if len(AG) >= 6 else AG
    ag_left11 = AG[:11] if len(AG) >= 11 else AG
    
    if ag_left4 == "NC07":
        return "알박구형"
    elif ag_left4 in ["23NA", "22NA", "24NA", "25NA"]:
        return "알박신형"
    elif "3J10" in AD:
        return "10kW"
    elif ag_left11 == "EVL-1C-22CQ":
        return "신형대"
    elif ag_left6 == "EVL-1C":
        return "구형대"
    elif ag_left4 == "EVL-" and "1107" in AD:
        return "신형대"
    elif ag_left4 == "EVL-":
        return "구형대"
    elif ag_left4 == "SBDA":
        return "신형대"
    elif ag_left4 == "SBAA":
        return "신형소"
    elif ag_left4 == "SBPA":
        return "F01" if "F01" in AD else "PC01"
    elif ag_left4 == "SBUA":
        return "UC01"
    elif ag_left4 == "SVI0":
        return "스필_7kW"
    elif ag_left3 == "E0C" or "CP" in AD:
        return "이카플러그"
    elif ag_left4 in ["1907", "1912"]:
        return "중앙제어_7kW"
    elif ag_left4 == "SC-P":
        return "SK_7kW"
    elif ag_left4 == "SANA":
        return "3kW"
    elif ag_left4 in ["EVS-", "007S"]:
        return "PNE_7kW"
    elif ag_left4 == "SBOA":
        return "F01" if "F01" in AD else "PC01"
    else:
        return "기타"

def create_sample_data():
    """초기 샘플 데이터 생성 (30개) - 실제 좌표 포함"""
    sample_data = {
        '사이트ID': [f'SITE_{i:03d}' for i in range(1, 21)] + [f'SITE_{i:03d}' for i in range(1, 11)],
        '모델분류': [
            '급속스필_100', '급속PNE_100', '신형대', '알박신형', '구형대',
            '급속SK_100', 'F01', '스필_7kW', '급속그린파워_100', 'PNE_7kW',
            '급속코스텔_50', '신형소', '이카플러그', '급속애플망고_200', 'UC01',
            '급속PNE_50', '10kW', 'SK_7kW', '중앙제어_7kW', '3kW',
            '급속스필_100', '신형대', '알박신형', '급속PNE_100', '구형대',
            '급속SK_200', 'F01', '급속그린파워_50', 'PNE_7kW', '급속알박_50'
        ],
        '권역': [
            '수도권북서', '수도권남동', '수도권북동', '충청권', '경상권',
            '수도권남서', '강원권', '수도권북서', '전라권', '수도권남동',
            '수도권북동', '충청권', '수도권남서', '경상권', '강원권',
            '수도권북서', '전라권', '수도권남동', '수도권북동', '충청권',
            '수도권북서', '수도권남동', '경상권', '강원권', '수도권북서',
            '전라권', '수도권남동', '수도권북동', '충청권', '수도권남서'
        ],
        '주소': [
            '서울특별시 강서구 공항대로 지하 396', '경기도 성남시 분당구 판교역로 166', '경기도 의정부시 상금로 36', '대전광역시 유성구 대학로 99', '부산광역시 해운대구 센텀중앙로 79',
            '인천광역시 계양구 안남로 560', '강원도 춘천시 중앙로 1', '서울특별시 마포구 월드컵북로 396', '광주광역시 서구 상무민주로 61', '경기도 용인시 수지구 포은대로 435',
            '서울특별시 강북구 4.19로 100', '충청남도 천안시 동남구 병천면 충절로 1600', '인천광역시 계양구 봉오대로744번길 7', '대구광역시 달서구 달구벌대로 1095', '강원도 원주시 세계로 123',
            '서울특별시 은평구 통일로 684', '전라북도 전주시 완산구 효자로 225', '경기도 수원시 팔달구 중부대로 120', '서울특별시 중랑구 망우로 379', '세종특별자치시 한누리대로 2130',
            '서울특별시 강서구 공항대로 지하 396', '경기도 성남시 분당구 판교역로 166', '울산광역시 남구 삼산로 282', '강원도 강릉시 경강로 2021', '서울특별시 양천구 목동서로 159',
            '전라남도 목포시 평화로 32', '경기도 성남시 중원구 사기막골로 45번길 14', '서울특별시 성북구 정릉로 77', '충청북도 청주시 상당구 상당로 82', '경상남도 창원시 의창구 중앙대로 151'
        ],
        '위도': [  # AN열 (40번째)
            37.5583, 37.3945, 37.7388, 36.3704, 35.1681,
            37.5376, 37.8813, 37.5665, 35.1595, 37.3217,
            37.6398, 36.5760, 37.5420, 35.8285, 37.3422,
            37.6176, 35.8242, 37.2636, 37.5985, 36.4801,
            37.5583, 37.3945, 35.5384, 37.7519, 37.5172,
            34.7943, 37.4201, 37.5894, 36.6424, 35.2272
        ],
        '경도': [  # AM열 (39번째)
            126.7944, 127.1116, 127.0467, 127.3622, 129.1303,
            126.7253, 127.7298, 126.9018, 126.8526, 127.1085,
            127.0253, 127.1472, 126.7389, 128.5658, 127.9202,
            126.9227, 127.1530, 127.0286, 127.0927, 127.2890,
            126.7944, 127.1116, 129.3114, 128.8761, 126.8664,
            126.3822, 127.1266, 127.0167, 127.4890, 128.6811
        ],
        '운영계약시작일': [
            date(2022, 1, 15), date(2022, 3, 20), date(2022, 5, 10), date(2022, 7, 5), date(2022, 9, 1),
            date(2023, 1, 10), date(2023, 3, 15), date(2023, 5, 20), date(2023, 7, 8), date(2023, 9, 12),
            date(2024, 1, 5), date(2024, 3, 10), date(2024, 5, 15), date(2024, 7, 20), date(2024, 9, 5),
            date(2025, 1, 8), date(2025, 3, 12), date(2025, 5, 18), date(2025, 7, 22), date(2025, 9, 10),
            date(2022, 2, 14), date(2022, 6, 18), date(2022, 10, 22), date(2023, 2, 15), date(2023, 6, 20),
            date(2024, 2, 10), date(2024, 6, 15), date(2024, 10, 20), date(2025, 2, 12), date(2025, 6, 18)
        ],
        '운영계약종료일': [
            date(2028, 1, 14), date(2028, 3, 19), date(2028, 5, 9), date(2028, 7, 4), date(2028, 8, 31),
            date(2029, 1, 9), date(2029, 3, 14), date(2029, 5, 19), date(2029, 7, 7), date(2029, 9, 11),
            date(2030, 1, 4), date(2030, 3, 9), date(2030, 5, 14), date(2030, 7, 19), date(2030, 9, 4),
            date(2031, 1, 7), date(2031, 3, 11), date(2031, 5, 17), date(2031, 7, 21), date(2031, 9, 9),
            date(2028, 2, 13), date(2028, 6, 17), date(2028, 10, 21), date(2029, 2, 14), date(2029, 6, 19),
            date(2030, 2, 9), date(2030, 6, 14), date(2030, 10, 19), date(2031, 2, 11), date(2031, 6, 17)
        ]
    }
    
    df = pd.DataFrame(sample_data)
    df['운영계약시작일_parsed'] = df['운영계약시작일']
    df['운영계약종료일_parsed'] = df['운영계약종료일']
    df['운영계약시작일_cleaned'] = df['운영계약시작일']
    df['운영계약종료일_cleaned'] = df['운영계약종료일']
    df['행번호'] = range(5, 5 + len(df))
    
    return df

def create_charger_map(filtered_df):
    """
    ??? 핵심 기능: 사이트ID로 그룹화된 충전기 지도 생성
    - AM열(경도), AN열(위도) 사용
    - 사이트ID별 1개 마커 표시
    - 네이버 지도 연동
    """
    # 유효한 좌표가 있는 데이터만 사용
    map_data = filtered_df.dropna(subset=['위도', '경도']).copy()
    
    if len(map_data) == 0:
        return None, "좌표 데이터가 없습니다. AM열(경도), AN열(위도)를 확인해주세요."
    
    # 사이트ID가 없는 경우 주소 기반으로 임시 생성
    if '사이트ID' not in map_data.columns or map_data['사이트ID'].isna().all():
        map_data['사이트ID'] = [f'SITE_{i:04d}' for i in range(len(map_data))]
    
    # ?? 핵심: 사이트ID로 그룹화 (같은 사이트의 여러 충전기를 1개 마커로)
    grouped = map_data.groupby('사이트ID').agg({
        '위도': 'first',
        '경도': 'first',
        '주소': 'first',
        '권역': 'first',
        '모델분류': lambda x: list(x),  # 모든 모델을 리스트로 수집
        '운영계약시작일': 'first',
        '운영계약종료일': 'first'
    }).reset_index()
    
    # 각 사이트의 충전기 개수 및 급속/완속 비율 계산
    charger_counts = map_data.groupby('사이트ID').agg({
        '모델분류': ['count', lambda x: sum('급속' in str(model) for model in x)]
    }).reset_index()
    
    charger_counts.columns = ['사이트ID', '총충전기수', '급속충전기수']
    charger_counts['완속충전기수'] = charger_counts['총충전기수'] - charger_counts['급속충전기수']
    
    # 데이터 병합
    grouped = grouped.merge(charger_counts, on='사이트ID')
    
    # 지도 중심점 계산
    center_lat = grouped['위도'].mean()
    center_lon = grouped['경도'].mean()
    
    # Folium 지도 생성
    m = folium.Map(
        location=[center_lat, center_lon],
        zoom_start=8,
        tiles='OpenStreetMap'
    )
    
    # 마커 클러스터링 추가
    marker_cluster = MarkerCluster(
        name="충전소 클러스터",
        overlay=True,
        control=True,
        options={
            "disableClusteringAtZoom": 15,
            "maxClusterRadius": 50
        }
    ).add_to(m)
    
    # 권역별 색상 매핑
    region_colors = {
        '수도권북서': 'blue', '수도권북동': 'green', '수도권남동': 'red', '수도권남서': 'purple',
        '수도권기타': 'cadetblue', '인천??': 'orange', '강원권': 'lightblue', '충청권': 'lightgreen',
        '경상권': 'pink', '전라권': 'lightgray', '기타': 'gray'
    }
    
    # 각 사이트별 마커 추가
    for idx, row in grouped.iterrows():
        site_id = row['사이트ID']
        address = row['주소']
        total_chargers = row['총충전기수']
        fast_chargers = row['급속충전기수']
        slow_chargers = row['완속충전기수']
        models = row['모델분류']
        region = row['권역']
        
        # ?? 네이버 지도 URL 생성 (주소 + " 전기차")
        encoded_address = quote(f"{address} 전기차")
        naver_map_url = f"https://map.naver.com/p/search/{encoded_address}"
        
        # 아이콘 설정 (급속 충전기가 있으면 번개, 없으면 플러그)
        icon_name = 'flash' if fast_chargers > 0 else 'plug'
        color = region_colors.get(region, 'gray')
        
        # ?? 마우스 오버 툴팁 (간단한 정보)
        tooltip_text = f"{site_id} | {total_chargers}기 | {region}"
        
        # ?? 클릭 시 팝업 (상세 정보 + 네이버 지도 링크)
        models_text = ', '.join(set([str(m) for m in models]))  # 중복 제거
        
        popup_html = f"""
        <div style="width: 320px; font-family: 'Malgun Gothic', Arial, sans-serif;">
            <h4 style="margin: 0 0 12px 0; color: #333; border-bottom: 3px solid {color}; padding-bottom: 8px; font-size: 16px;">
                ?? {site_id}
            </h4>
            <table style="width: 100%; font-size: 13px; border-collapse: collapse; margin-bottom: 12px;">
                <tr style="background-color: #f8f9fa;">
                    <td style="padding: 6px; font-weight: bold; width: 80px;">총 충전기</td>
                    <td style="padding: 6px; color: #0066cc; font-weight: bold;">{total_chargers}대</td>
                </tr>
                <tr>
                    <td style="padding: 6px; font-weight: bold;">급속/완속</td>
                    <td style="padding: 6px;">? {fast_chargers}대 / ?? {slow_chargers}대</td>
                </tr>
                <tr style="background-color: #f8f9fa;">
                    <td style="padding: 6px; font-weight: bold;">권역</td>
                    <td style="padding: 6px;">{region}</td>
                </tr>
                <tr>
                    <td style="padding: 6px; font-weight: bold;">주소</td>
                    <td style="padding: 6px; font-size: 11px; line-height: 1.3;">{address}</td>
                </tr>
                <tr style="background-color: #f8f9fa;">
                    <td style="padding: 6px; font-weight: bold;">모델</td>
                    <td style="padding: 6px; font-size: 11px; line-height: 1.3;">{models_text}</td>
                </tr>
            </table>
            <div style="text-align: center; margin-top: 15px;">
                <a href="{naver_map_url}" target="_blank" style="
                    display: inline-block;
                    padding: 10px 20px;
                    background: linear-gradient(135deg, #03C75A 0%, #029B47 100%);
                    color: white;
                    text-decoration: none;
                    border-radius: 8px;
                    font-weight: bold;
                    font-size: 13px;
                    box-shadow: 0 3px 8px rgba(3, 199, 90, 0.3);
                    transition: all 0.2s ease;
                ">
                    ?? 네이버 지도에서 보기 ↗
                </a>
            </div>
        </div>
        """
        
        # 마커 추가
        folium.Marker(
            location=[row['위도'], row['경도']],
            popup=folium.Popup(popup_html, max_width=350),
            tooltip=tooltip_text,
            icon=folium.Icon(
                color=color,
                icon=icon_name,
                prefix='fa'
            )
        ).add_to(marker_cluster)
    
    # ?? 범례 추가
    legend_html = f'''
    <div style="position: fixed; bottom: 50px; right: 50px; border: 2px solid grey; 
                z-index: 9999; background-color: white; padding: 15px; font-size: 12px;
                border-radius: 8px; box-shadow: 0 4px 12px rgba(0,0,0,0.3); font-family: Arial;">
        <p style="margin: 0 0 10px 0; font-weight: bold; font-size: 14px; color: #333;">??? 지도 범례</p>
        <p style="margin: 5px 0;"><i class="fa fa-flash" style="color: red;"></i> 급속 충전기 포함 사이트</p>
        <p style="margin: 5px 0;"><i class="fa fa-plug" style="color: blue;"></i> 완속 충전기만 있는 사이트</p>
        <hr style="margin: 8px 0;">
        <p style="margin: 5px 0; font-weight: bold;">총 사이트: {len(grouped):,}개</p>
        <p style="margin: 5px 0; font-weight: bold;">총 충전기: {len(map_data):,}대</p>
        <hr style="margin: 8px 0;">
        <p style="margin: 2px 0; font-size: 11px; color: #666;">권역별 색상:</p>
    '''
    
    # 권역별 색상 정보 추가
    for region, color in region_colors.items():
        if region in grouped['권역'].values:
            site_count = len(grouped[grouped['권역'] == region])
            legend_html += f'<p style="margin: 1px 0; font-size: 10px;"><span style="color: {color}; font-size: 14px;">●</span> {region} ({site_count}개)</p>'
    
    legend_html += '</div>'
    m.get_root().html.add_child(folium.Element(legend_html))
    
    return m, None

def process_excel_file_with_progress(file_bytes, title_container, progress_bar, status_text):
    """실시간 타이머와 함께 엑셀 파일을 처리하고 대시보드용 DataFrame 생성"""
    try:
        start_time = time.time()
        
        # 1단계: 파일 로드
        elapsed = time.time() - start_time
        title_container.markdown(f"### ?? 작업 진행 상황 `?? {format_time(elapsed)}`")
        status_text.markdown("?? **엑셀 파일을 읽는 중입니다...**")
        progress_bar.progress(5)
        
        file_stream = io.BytesIO(file_bytes)
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        ws = wb.active
        
        # 열 번호 정의
        AR_COLUMN = 44  # 운영계약 시작일
        AS_COLUMN = 45  # 운영계약 종료일
        AM_COLUMN = 39  # 경도
        AN_COLUMN = 40  # 위도
        BA_COLUMN = 53  # 모델분류
        BB_COLUMN = 54  # 권역
        
        # 헤더 추가
        ws.cell(row=4, column=BA_COLUMN, value='모델분류')
        ws.cell(row=4, column=BB_COLUMN, value='권역')
        
        # 2단계: 데이터 분석
        elapsed = time.time() - start_time
        title_container.markdown(f"### ?? 작업 진행 상황 `?? {format_time(elapsed)}`")
        status_text.markdown("?? **데이터 구조를 분석하는 중입니다...**")
        progress_bar.progress(15)
        
        max_row = ws.max_row
        if max_row < 5:
            max_row = 5
        
        max_col = max(ws.max_column, 54)
        total_rows = max_row - 4
        
        dashboard_data = []
        ar_cleaned_count = 0
        as_cleaned_count = 0
        
        # 3단계: 분류 및 날짜 정리 작업
        status_text.markdown(f"? **모델분류, 권역분류 및 날짜 정리 작업을 시작합니다... (총 {total_rows:,}개 행)**")
        progress_bar.progress(20)
        
        processed_count = 0
        last_update_time = time.time()
        update_interval = 0.5
        
        for i, row_num in enumerate(range(5, max_row + 1)):
            row_data = {}
            for col_num in range(1, max_col + 1):
                col_letter = get_column_letter(col_num)
                cell_value = ws.cell(row=row_num, column=col_num).value
                row_data[col_letter] = cell_value
            
            # BA열: 모델 분류
            classification_result = classify_model(row_data, row_num)
            ws.cell(row=row_num, column=BA_COLUMN, value=classification_result)
            
            # BB열: 권역 분류
            address = get_safe_value(row_data, 'H')
            region_result = classify_region(address)
            ws.cell(row=row_num, column=BB_COLUMN, value=region_result)
            
            # AR열 날짜 정리
            ar_value = row_data.get('AR')
            ar_cleaned = clean_and_parse_date(ar_value)
            if ar_cleaned:
                ar_formatted = format_date_for_excel(ar_cleaned)
                ws.cell(row=row_num, column=AR_COLUMN, value=ar_formatted)
                ws.cell(row=row_num, column=AR_COLUMN).number_format = 'YYYY-MM-DD'
                ar_cleaned_count += 1
            
            # AS열 날짜 정리
            as_value = row_data.get('AS')
            as_cleaned = clean_and_parse_date(as_value)
            if as_cleaned:
                as_formatted = format_date_for_excel(as_cleaned)
                ws.cell(row=row_num, column=AS_COLUMN, value=as_formatted)
                ws.cell(row=row_num, column=AS_COLUMN).number_format = 'YYYY-MM-DD'
                as_cleaned_count += 1
            
            # ?? 좌표 데이터 수집 (AM열=경도, AN열=위도)
            site_id = get_safe_value(row_data, 'A')  # A열을 사이트ID로 가정
            longitude = row_data.get('AM')  # 경도
            latitude = row_data.get('AN')   # 위도
            
            # 숫자로 변환 시도
            try:
                longitude = float(longitude) if longitude else None
            except:
                longitude = None
            
            try:
                latitude = float(latitude) if latitude else None
            except:
                latitude = None
            
            dashboard_data.append({
                '사이트ID': site_id if site_id else f'AUTO_{row_num}',
                '모델분류': classification_result,
                '권역': region_result,
                '주소': address,
                '위도': latitude,
                '경도': longitude,
                '운영계약시작일': ar_value,
                '운영계약종료일': as_value,
                '운영계약시작일_cleaned': ar_cleaned,
                '운영계약종료일_cleaned': as_cleaned,
                '행번호': row_num
            })
            
            processed_count += 1
            
            current_time = time.time()
            should_update = (
                (i + 1) % 50 == 0 or
                (current_time - last_update_time) >= update_interval or
                i == total_rows - 1
            )
            
            if should_update:
                elapsed_time = current_time - start_time
                progress_percent = 20 + int((processed_count / total_rows) * 70)
                progress_bar.progress(progress_percent)
                
                title_container.markdown(f"### ?? 작업 진행 상황 `?? {format_time(elapsed_time)}`")
                
                if processed_count > 0:
                    rows_per_second = processed_count / elapsed_time
                    remaining_rows = total_rows - processed_count
                    estimated_remaining = remaining_rows / rows_per_second if rows_per_second > 0 else 0
                    
                    status_text.markdown(
                        f"? **분류 진행 중...** `{processed_count:,}/{total_rows:,}` 완료 "
                        f"({(processed_count/total_rows*100):.1f}%) | "
                        f"?? **처리 속도:** `{rows_per_second:.1f}행/초` | "
                        f"? **예상 남은 시간:** `{format_time(estimated_remaining)}`"
                    )
                
                last_update_time = current_time
        
        # 4단계: 파일 저장
        elapsed = time.time() - start_time
        title_container.markdown(f"### ?? 작업 진행 상황 `?? {format_time(elapsed)}`")
        status_text.markdown("?? **결과 파일을 생성하는 중입니다...**")
        progress_bar.progress(95)
        
        output_stream = io.BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)
        
        df = pd.DataFrame(dashboard_data)
        df['운영계약시작일_parsed'] = df['운영계약시작일_cleaned']
        df['운영계약종료일_parsed'] = df['운영계약종료일_cleaned']
        
        total_time = time.time() - start_time
        progress_bar.progress(100)
        
        title_container.markdown(f"### ?? 작업 완료! `? 총 {format_time(total_time)}`")
        
        avg_speed = processed_count / total_time if total_time > 0 else 0
        status_text.markdown(
            f"?? **처리 완료!** `{processed_count:,}개 행`이 성공적으로 분류되었습니다. | "
            f"**평균 속도:** `{avg_speed:.1f}행/초` | "
            f"?? **날짜 정리:** AR열 `{ar_cleaned_count:,}개`, AS열 `{as_cleaned_count:,}개`"
        )
        
        return output_stream, None, processed_count, total_time, df, ar_cleaned_count, as_cleaned_count
        
    except Exception as e:
        import traceback
        elapsed = time.time() - start_time
        title_container.markdown(f"### ? 작업 중단 `?? {format_time(elapsed)}`")
        status_text.markdown("? **오류가 발생했습니다.**")
        error_detail = traceback.format_exc()
        return None, f"파일 처리 중 오류 발생: {str(e)}\n\n상세:\n{error_detail}", 0, 0, None, 0, 0

def show_dashboard(df):
    """대시보드 화면을 표시하는 함수"""
    st.markdown("## ?? 충전기 운영 현황 대시보드")
    
    # 날짜 필터 섹션
    st.markdown("### ??? 운영계약 기간 필터")
    
    valid_dates = df.dropna(subset=['운영계약시작일_parsed', '운영계약종료일_parsed'])
    
    if len(valid_dates) > 0:
        # pd.to_datetime을 통해 안전하게 변환 후 순수 date 객체로 뽑아냅니다.
        min_date = pd.to_datetime(valid_dates['운영계약시작일_parsed']).min().date()
        max_date = pd.to_datetime(valid_dates['운영계약종료일_parsed']).max().date()
        
        # 안전한 기본값 계산
        default_start = max(min_date, date(2022, 1, 1))
        default_end = min(max_date, date(2028, 1, 1))
        
        if default_start > default_end:
            default_start = min_date
            default_end = max_date
        
        col1, col2, col3 = st.columns([2, 2, 1])
        
        with col1:
            start_date = st.date_input(
                "계약 시작일 (이후)",
                value=default_start,
                min_value=min_date,
                max_value=max_date,
                help="AR열 기준 - 이 날짜 이후에 시작하는 계약"
            )
        
        with col2:
            end_date = st.date_input(
                "계약 종료일 (이전)",
                value=default_end,
                min_value=min_date,
                max_value=max_date,
                help="AS열 기준 - 이 날짜 이전에 종료하는 계약"
            )
        
        with col3:
            st.markdown("<br>", unsafe_allow_html=True)
            filter_applied = st.button("?? 필터 적용", type="primary", use_container_width=True)
        
        mask = (
            (df['운영계약시작일_parsed'] < end_date) & 
            (df['운영계약종료일_parsed'] >= start_date) &
            df['운영계약시작일_parsed'].notna() &
            df['운영계약종료일_parsed'].notna()
        )
        
        filtered_df = df[mask].copy()
        
        st.info(
            f"?? **선택 기간:** {start_date} ~ {end_date} | "
            f"**해당 기간 충전기:** {len(filtered_df):,}대 (전체 {len(df):,}대 중 {len(filtered_df)/len(df)*100:.1f}%)"
        )
        
        if len(filtered_df) == 0:
            st.warning("?? 선택한 기간에 해당하는 데이터가 없습니다. 필터 조건을 조정해주세요.")
            return
        
        st.markdown("---")
        
        # 1행: 주요 지표 (?? 여기서 모든 변수를 미리 계산)
        st.markdown("### ?? 주요 지표")
        
        # ? 핵심 수정: 모든 변수를 여기서 미리 계산
        total_chargers = len(filtered_df)
        unique_sites = filtered_df['사이트ID'].nunique() if '사이트ID' in filtered_df.columns else 0
        region_count = filtered_df['권역'].nunique()
        model_count = filtered_df['모델분류'].nunique()  # ?? 오류 해결: 여기서 계산
        fast_chargers = len(filtered_df[filtered_df['모델분류'].str.contains('급속', na=False)])
        fast_ratio = (fast_chargers / total_chargers * 100) if total_chargers > 0 else 0
        
        kpi_col1, kpi_col2, kpi_col3, kpi_col4 = st.columns(4)
        
        with kpi_col1:
            st.metric("총 충전기 수", f"{total_chargers:,}대")
        
        with kpi_col2:
            st.metric("사이트 수", f"{unique_sites:,}개")
        
        with kpi_col3:
            st.metric("권역 수", f"{region_count}개")
        
        with kpi_col4:
            st.metric("급속 충전기", f"{fast_chargers:,}대", f"{fast_ratio:.1f}%")
        
        st.markdown("---")
        
        # ??? 지도 시각화 (최상단 배치)
        st.markdown("### ??? 충전기 위치 지도 (사이트ID 기준)")
        
        # 좌표 데이터 확인
        has_coordinates = '위도' in filtered_df.columns and '경도' in filtered_df.columns
        
        if has_coordinates:
            valid_coords = filtered_df.dropna(subset=['위도', '경도'])
            coord_count = len(valid_coords)
            
            if coord_count > 0:
                # 사이트 수 계산
                unique_sites_map = valid_coords['사이트ID'].nunique() if '사이트ID' in valid_coords.columns else coord_count
                
                st.success(f"? {unique_sites_map:,}개 사이트, {coord_count:,}개 충전기의 좌표 데이터가 있습니다.")
                
                # 지도 생성
                charger_map, error = create_charger_map(filtered_df)
                
                if error:
                    st.error(f"? {error}")
                else:
                    st_folium(charger_map, width=1400, height=700)
                    
                    # 지도 통계
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("지도 표시 사이트", f"{unique_sites_map:,}개")
                    with col2:
                        st.metric("지도 표시 충전기", f"{coord_count:,}대")
                    with col3:
                        coord_ratio = (coord_count / len(filtered_df) * 100) if len(filtered_df) > 0 else 0
                        st.metric("좌표 보유율", f"{coord_ratio:.1f}%")
                    with col4:
                        missing_coords = len(filtered_df) - coord_count
                        st.metric("좌표 누락", f"{missing_coords:,}대")
                    
                    # 사용법 안내
                    with st.expander("?? 지도 사용 방법"):
                        st.markdown("""
                        **지도 조작:**
                        - **확대/축소:** 마우스 휠 또는 +/- 버튼
                        - **이동:** 드래그
                        - **클러스터:** 숫자 표시된 원 클릭 시 해당 영역 확대
                        
                        **마커 기능:**
                        - ??? **마우스 오버:** 사이트ID, 충전기 수, 권역 간단 정보
                        - ??? **클릭:** 상세 정보 팝업 + 네이버 지도 링크
                        - ?? **네이버 지도 버튼:** 클릭 시 새 창에서 해당 위치의 전기차 충전소 검색
                        
                        **시각화 정보:**
                        - ? **빨간 번개:** 급속 충전기 포함 사이트
                        - ?? **파란 플러그:** 완속 충전기만 있는 사이트
                        - ?? **색상:** 권역별 구분 (범례 참조)
                        - ?? **그룹화:** 같은 사이트ID는 1개 마커로 표시
                        """)
            else:
                st.warning("?? 좌표 데이터가 없습니다. AM열(경도), AN열(위도)를 확인해주세요.")
        else:
            st.warning("?? 좌표 데이터가 없습니다. AM열(경도), AN열(위도)를 엑셀 파일에 추가해주세요.")
        
        st.markdown("---")
        
        # 2행: 모델별 현황
        st.markdown("### ? 모델별 충전기 현황")
        
        col1, col2 = st.columns([3, 2])
        
        with col1:
            model_counts = filtered_df['모델분류'].value_counts().reset_index()
            model_counts.columns = ['모델분류', '수량']
            
            fig_model = px.bar(
                model_counts.head(15),
                x='수량',
                y='모델분류',
                orientation='h',
                title='모델별 충전기 수량 (상위 15개)',
                labels={'수량': '충전기 수량 (대)', '모델분류': '모델'},
                color='수량',
                color_continuous_scale='Blues',
                text='수량'
            )
            fig_model.update_layout(height=500, showlegend=False)
            fig_model.update_traces(texttemplate='%{text}', textposition='outside')
            st.plotly_chart(fig_model, use_container_width=True)
        
        with col2:
            st.markdown("#### ?? 모델별 수량 상세")
            model_counts['비율'] = (model_counts['수량'] / model_counts['수량'].sum() * 100).round(1)
            model_counts['비율'] = model_counts['비율'].astype(str) + '%'
            
            st.dataframe(
                model_counts[['모델분류', '수량', '비율']],
                use_container_width=True,
                hide_index=True,
                height=450
            )
        
        st.markdown("---")
        
        # 3행: 권역별 현황
        st.markdown("### ??? 권역별 충전기 현황")
        
        col1, col2 = st.columns([2, 3])
        
        with col1:
            region_counts = filtered_df['권역'].value_counts().reset_index()
            region_counts.columns = ['권역', '수량']
            
            fig_region_pie = px.pie(
                region_counts,
                values='수량',
                names='권역',
                title='권역별 충전기 비율',
                hole=0.4
            )
            fig_region_pie.update_traces(textposition='inside', textinfo='percent+label')
            fig_region_pie.update_layout(height=400)
            st.plotly_chart(fig_region_pie, use_container_width=True)
        
        with col2:
            fig_region_bar = px.bar(
                region_counts,
                x='권역',
                y='수량',
                title='권역별 충전기 수량',
                labels={'수량': '충전기 수량 (대)', '권역': '권역'},
                color='수량',
                color_continuous_scale='Greens',
                text='수량'
            )
            fig_region_bar.update_layout(height=400, showlegend=False)
            fig_region_bar.update_traces(texttemplate='%{text}', textposition='outside')
            st.plotly_chart(fig_region_bar, use_container_width=True)
        
        st.markdown("---")
        
        # 4행: 권역별 모델 분포 히트맵
        st.markdown("### ?? 권역별 모델 분포 히트맵")
        
        crosstab = pd.crosstab(filtered_df['권역'], filtered_df['모델분류'])
        top_models = filtered_df['모델분류'].value_counts().head(12).index
        crosstab_filtered = crosstab[top_models]
        
        fig_heatmap = px.imshow(
            crosstab_filtered.T,
            labels=dict(x="권역", y="모델분류", color="수량"),
            x=crosstab_filtered.index,
            y=crosstab_filtered.columns,
            color_continuous_scale='RdYlGn',
            aspect="auto",
            title='권역별 주요 모델 분포 (상위 12개 모델)',
            text_auto=True
        )
        fig_heatmap.update_layout(height=500)
        st.plotly_chart(fig_heatmap, use_container_width=True)
        
        st.markdown("---")
        
        # 5행: 상세 데이터 테이블
        st.markdown("### ?? 권역별 × 모델별 상세 현황")
        
        pivot_wide = pd.crosstab(filtered_df['권역'], filtered_df['모델분류'], margins=True)
        
        st.dataframe(
            pivot_wide,
            use_container_width=True,
            height=400
        )
        
        # 다운로드 섹션
        st.markdown("---")
        st.markdown("### ?? 데이터 다운로드")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            csv_pivot = pivot_wide.to_csv(encoding='utf-8-sig')
            st.download_button(
                label="?? 권역×모델 현황표 CSV",
                data=csv_pivot,
                file_name=f"권역모델현황_{start_date}_{end_date}.csv",
                mime="text/csv",
                use_container_width=True
            )
        
        with col2:
            csv_filtered = filtered_df.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="?? 필터링된 전체 데이터 CSV",
                data=csv_filtered,
                file_name=f"필터링데이터_{start_date}_{end_date}.csv",
                mime="text/csv",
                use_container_width=True
            )
        
        with col3:
            # ? 이제 model_count가 정의되어 있으므로 정상 작동
            summary_report = f"""충전기 현황 요약 리포트

?? 분석 기간: {start_date} ~ {end_date}
?? 총 충전기 수: {total_chargers:,}대
?? 총 사이트 수: {unique_sites:,}개
?? 모델 종류: {model_count}개
??? 권역 수: {region_count}개
? 급속 충전기: {fast_chargers:,}대 ({fast_ratio:.1f}%)

?? 상위 5개 모델:
{chr(10).join([f"{i+1}. {row['모델분류']}: {row['수량']:,}대 ({row['비율']})" for i, row in model_counts.head(5).iterrows()])}

??? 권역별 분포:
{chr(10).join([f"? {row['권역']}: {row['수량']:,}대" for _, row in region_counts.iterrows()])}
"""
            
            st.download_button(
                label="?? 요약 리포트 TXT",
                data=summary_report,
                file_name=f"요약리포트_{start_date}_{end_date}.txt",
                mime="text/plain",
                use_container_width=True
            )
        
        # 데이터 품질 체크 섹션
        st.markdown("---")
        st.markdown("### ?? 데이터 품질 체크")
        
        problematic_regions = ['수도권??', '인천??', '기타', '수도권기타']
        unknown_regions = filtered_df[
            filtered_df['권역'].isin(problematic_regions)
        ]
        
        col1, col2 = st.columns(2)
        
        with col1:
            normal_count = len(filtered_df) - len(unknown_regions)
            normal_ratio = (normal_count / len(filtered_df) * 100) if len(filtered_df) > 0 else 0
            st.metric("정상 분류", f"{normal_count:,}대", f"{normal_ratio:.1f}%")
        
        with col2:
            unknown_ratio = (len(unknown_regions) / len(filtered_df) * 100) if len(filtered_df) > 0 else 0
            st.metric("미분류/불명확", f"{len(unknown_regions):,}대", f"{unknown_ratio:.1f}%")
        
        if len(unknown_regions) > 0:
            st.warning(f"?? {len(unknown_regions):,}개의 주소가 미분류되었거나 불명확합니다.")
            
            with st.expander("?? 미분류 주소 상세 보기", expanded=False):
                unknown_stats = unknown_regions['권역'].value_counts()
                st.markdown("**권역별 미분류 통계:**")
                st.dataframe(
                    unknown_stats.reset_index().rename(columns={'권역': '권역', 'count': '수량'}),
                    use_container_width=True,
                    hide_index=True
                )
                
                st.markdown("**상세 주소 목록 (최대 10개):**")
                display_cols = ['주소', '권역', '모델분류', '사이트ID']
                available_cols = [col for col in display_cols if col in unknown_regions.columns]
                
                st.dataframe(
                    unknown_regions[available_cols].head(10),
                    use_container_width=True,
                    hide_index=True
                )
                
                if len(unknown_regions) > 10:
                    st.info(f"?? 총 {len(unknown_regions):,}개 중 10개만 표시됩니다.")
        else:
            st.success("? 모든 주소가 정확하게 분류되었습니다!")
    
    else:
        st.warning("?? 유효한 운영계약 날짜 데이터가 없습니다. AR열과 AS열을 확인해주세요.")

def main():
    # 세션 상태 초기화 시 샘플 데이터 자동 로드
    if 'processed_df' not in st.session_state:
        st.session_state.processed_df = create_sample_data()
        st.session_state.is_sample_data = True
    if 'processed_file' not in st.session_state:
        st.session_state.processed_file = None
    
    # 헤더
    st.title("? 충전기 모델분류 & 운영현황 대시보드")
    st.markdown("""
    엑셀 파일을 업로드하면 **BA열**에 "모델분류", **BB열**에 "권역"을 자동으로 추가하고,  
    **AR열/AS열**의 날짜 데이터를 정리하여 **AM열(경도)/AN열(위도)** 기반으로 지도에서 충전기 위치를 확인할 수 있습니다.
    """)
    
    # 탭 구성
    tab1, tab2 = st.tabs(["?? 파일 업로드 & 분류", "?? 운영현황 대시보드"])
    
    with tab1:
        # 샘플 데이터 안내
        if st.session_state.get('is_sample_data', False):
            st.info("?? **현재 샘플 데이터가 로드되어 있습니다.** '운영현황 대시보드' 탭에서 바로 지도 기능을 체험해보세요!")
        
        uploaded_file = st.file_uploader(
            "?? 엑셀 파일을 선택하세요",
            type=['xlsx', 'xls'],
            help="AM열(경도), AN열(위도), 사이트ID가 포함된 파일을 권장합니다. 최대 200MB"
        )
        
        if uploaded_file is not None:
            col1, col2 = st.columns([3, 1])
            with col1:
                st.info(f"?? **{uploaded_file.name}**")
            with col2:
                file_size_mb = uploaded_file.size / (1024 * 1024)
                if file_size_mb >= 1:
                    st.metric("파일 크기", f"{file_size_mb:.1f} MB")
                else:
                    st.metric("파일 크기", f"{uploaded_file.size / 1024:.1f} KB")
            
            if st.button("?? 모델분류 시작", type="primary", use_container_width=True):
                
                title_container = st.empty()
                title_container.markdown("### ?? 작업 진행 상황 `?? 0.0초`")
                
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                st.markdown("---")
                
                file_bytes = uploaded_file.read()
                result = process_excel_file_with_progress(
                    file_bytes, 
                    title_container,
                    progress_bar, 
                    status_text
                )
                
                processed_file, error, processed_count, total_time, df, ar_cleaned, as_cleaned = result
                
                if error:
                    st.error(f"? {error}")
                else:
                    # 샘플 데이터 플래그 제거
                    st.session_state.processed_df = df
                    st.session_state.processed_file = processed_file
                    st.session_state.is_sample_data = False
                    
                    st.success(
                        f"?? **축하합니다!** 총 **{processed_count:,}개 행**의 모델분류 및 권역분류가 "
                        f"**{format_time(total_time)}**만에 완료되었습니다!"
                    )
                    
                    st.info(
                        f"?? **날짜 정리 완료:** AR열(운영계약시작일) `{ar_cleaned:,}개`, "
                        f"AS열(운영계약종료일) `{as_cleaned:,}개` - 시간 부분(00:00:00) 제거 및 YYYY-MM-DD 형식으로 변환"
                    )
                    
                    # 한국 시간 적용
                    korea_time = get_korea_time()
                    timestamp = korea_time.strftime("%Y%m%d_%H%M%S")
                    download_name = f"모델분류_결과_{timestamp}.xlsx"
                    
                    col1, col2, col3 = st.columns([1, 2, 1])
                    with col2:
                        st.download_button(
                            label="?? 결과 파일 다운로드",
                            data=processed_file.getvalue(),
                            file_name=download_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            type="primary"
                        )
                    
                    st.info("?? **'운영현황 대시보드'** 탭으로 이동하여 지도와 데이터를 분석해보세요!")
                    
                    with st.expander("?? 처리 결과 상세 정보", expanded=False):
                        col1, col2, col3, col4, col5 = st.columns(5)
                        with col1:
                            st.metric("처리된 행 수", f"{processed_count:,}개")
                        with col2:
                            st.metric("소요 시간", format_time(total_time))
                        with col3:
                            avg_speed = processed_count / total_time if total_time > 0 else 0
                            st.metric("처리 속도", f"{avg_speed:.1f}행/초")
                        with col4:
                            st.metric("AR열 정리", f"{ar_cleaned:,}개")
                        with col5:
                            st.metric("AS열 정리", f"{as_cleaned:,}개")
        
        with st.expander("?? 분류 기준 정보 보기", expanded=False):
            show_classification_info()
    
    with tab2:
        if st.session_state.processed_df is not None:
            # 샘플 데이터 사용 중인 경우 알림
            if st.session_state.get('is_sample_data', False):
                st.warning("?? **현재 샘플 데이터를 사용 중입니다.** 실제 데이터를 분석하려면 '파일 업로드 & 분류' 탭에서 파일을 업로드하세요.")
            
            show_dashboard(st.session_state.processed_df)
        else:
            st.info("?? 먼저 **'파일 업로드 & 분류'** 탭에서 파일을 업로드하고 분류 작업을 완료해주세요.")

def show_classification_info():
    """분류 기준 정보를 표시하는 함수"""
    st.markdown("### ?? 모델분류 기준표")
    
    subtab1, subtab2, subtab3, subtab4 = st.tabs(["? 급속 충전기", "?? 완속 충전기", "??? 권역 분류", "?? 참조 정보"])
    
    with subtab1:
        st.markdown("#### ? 급속 충전기 분류 기준")
        st.info("**전제 조건:** AH열 = '급속'")
        
        fast_charger_data = {
            "모델분류명": [
                "급속스필_100", "급속스필_50", "급속PNE_100", "급속PNE_50",
                "급속PNE_200", "급속PNE_150", "급속애플망고_200", "급속SK_100",
                "급속SK_200", "급속코스텔_50", "급속중앙제어_50", "급속그린파워_100",
                "급속그린파워_50", "급속알박_50", "급속"
            ],
            "AG열 코드 조건": [
                "S0F1", "S0F5", "EVQ- (AJ=100)", "EVQ- 또는 EV1- (AJ=50)",
                "MAXE", "DP15", "A01- 또는 AD1-", "Q081, Q101, Q010",
                "Q071, Q102", "1Y25, 1Y24", "1911", "1900",
                "19C0", "QC50", "위 조건 해당 없음"
            ],
            "제조사": [
                "스필", "스필", "PNE", "PNE", "PNE", "PNE", "애플망고",
                "SK", "SK", "코스텔", "중앙제어", "그린파워", "그린파워",
                "알박", "기타 급속"
            ]
        }
        
        st.dataframe(fast_charger_data, use_container_width=True, hide_index=True)
    
    with subtab2:
        st.markdown("#### ?? 완속 충전기 분류 기준")
        st.info("**전제 조건:** AH열 ≠ '급속'")
        
        slow_charger_data = {
            "우선순위": ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10",
                        "11", "12", "13", "14", "15", "16", "17", "18", "19"],
            "모델분류명": [
                "알박구형", "알박신형", "10kW", "신형대 (EVL-1C-22CQ)", "구형대 (EVL-1C)",
                "신형대 (EVL+1107)", "구형대 (EVL 기본)", "신형대 (SBDA)", "신형소",
                "F01 / PC01 (SBPA)", "UC01", "스필_7kW", "이카플러그", "중앙제어_7kW",
                "SK_7kW", "3kW", "PNE_7kW", "F01 / PC01 (SBOA)", "기타"
            ]
        }
        
        st.dataframe(slow_charger_data, use_container_width=True, hide_index=True)
    
    with subtab3:
        st.markdown("#### ??? 권역 분류 기준 (H열 주소 기반)")
        
        region_data = {
            "권역명": ["수도권북서", "수도권북동", "수도권남동", "수도권남서", 
                      "수도권남서(인천)", "수도권기타", "강원권", "충청권", "경상권", "전라권", "기타"],
            "주요 지역": [
                "고양, 부천, 김포, 파주, 은평, 마포, 서대문, 양천, 강서",
                "도봉, 노원, 중랑, 강북, 성북, 동대문, 의정부, 남양주, 구리",
                "강남, 서초, 송파, 강동, 성남, 용인, 하남, 수원, 평택",
                "구로, 금천, 영등포, 동작, 관악, 의왕, 광명, 안산, 안양",
                "계양, 남동, 부평, 연수, 미추홀 등 인천 주요 구",
                "경기도 내 특정 권역 미분류 지역",
                "강원도 전역",
                "충청, 충남, 충북, 세종, 대전",
                "경상, 경남, 경북, 부산, 대구, 울산",
                "전라, 전남, 전북, 광주",
                "위 조건 해당 없음"
            ]
        }
        
        st.dataframe(region_data, use_container_width=True, hide_index=True)
    
    with subtab4:
        st.markdown("#### ?? 엑셀 열 참조 정보")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("""
            **?? 분류에 사용되는 엑셀 열**
            
            | 열 이름 | 열 번호 | 용도 |
            |---------|---------|------|
            | **E열** | 5번째 | 사이트ID (지도 그룹화) ?? |
            | **H열** | 8번째 | 주소 (권역 분류) |
            | **AD열** | 30번째 | 모델명/설명 검색 |
            | **AG열** | 33번째 | 모델 코드 (주요 기준) |
            | **AH열** | 34번째 | 급속/완속 구분 |
            | **AJ열** | 36번째 | 용량 정보 (kW) |
            | **AM열** | 39번째 | 경도 (Longitude) ??? |
            | **AN열** | 40번째 | 위도 (Latitude) ??? |
            | **AR열** | 44번째 | 운영계약 시작일 ?정리 |
            | **AS열** | 45번째 | 운영계약 종료일 ?정리 |
            """)
        
        with col2:
            st.markdown("""
            **?? 결과 출력 위치**
            
            | 항목 | 위치 | 내용 |
            |------|------|------|
            | **모델분류 헤더** | BA열 4행 | "모델분류" |
            | **권역 헤더** | BB열 4행 | "권역" |
            | **모델분류 결과** | BA열 5행~ | 각 행별 분류값 |
            | **권역 결과** | BB열 5행~ | 각 행별 권역값 |
            | **AR열 정리** | AR열 5행~ | YYYY-MM-DD 형식 |
            | **AS열 정리** | AS열 5행~ | YYYY-MM-DD 형식 |
            """)
        
        st.success("""
        **??? 지도 기능:**
        - **AM열(경도), AN열(위도)** 좌표로 정확한 위치 표시
        - **사이트ID(E열)** 기준으로 같은 위치 충전기 그룹화
        - **마우스 오버:** 사이트ID, 충전기 수, 권역 정보
        - **클릭:** 상세 정보 + 네이버 지도 링크
        - **네이버 지도:** "주소 + 전기차" 검색으로 새 창 열기
        """)

if __name__ == "__main__":
    main()
